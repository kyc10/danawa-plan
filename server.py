from flask import Flask, jsonify, request, send_from_directory
from flask_cors import CORS
import openpyxl
import json
import os
from datetime import datetime

app = Flask(__name__, static_folder='static')
CORS(app)

DATA_FILE = os.path.join(os.path.dirname(__file__), 'data.json')
UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

CATEGORIES = ['식료품', '꾸밈비', '외식비', '문화비', '의료비', '공과금', '교육비',
              '교통비', '보험료', '차량비용', '여행경비', '대출이자', '경조사비', '생필품', '기타']

def parse_excel(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    monthly_data = {}
    for sheet_name in wb.sheetnames:
        if len(sheet_name) == 4 and sheet_name.isdigit():
            yy = int(sheet_name[:2])
            mm = int(sheet_name[2:])
            if 1 <= mm <= 12:
                year = 2000 + yy
                month_key = f"{year}-{mm:02d}"
                ws = wb[sheet_name]
                month_cats = {}
                transactions = []
                for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
                    cat_name = row[1] if len(row) > 1 else None
                    cat_val = row[2] if len(row) > 2 else None
                    tx_date = row[5] if len(row) > 5 else None
                    tx_name = row[6] if len(row) > 6 else None
                    tx_amount = row[7] if len(row) > 7 else None
                    tx_category = row[9] if len(row) > 9 else None
                    if cat_name in CATEGORIES and isinstance(cat_val, (int, float)):
                        month_cats[cat_name] = cat_val
                    if tx_name and isinstance(tx_amount, (int, float)) and tx_category and tx_name != '-':
                        date_str = None
                        if isinstance(tx_date, datetime):
                            date_str = tx_date.strftime('%Y-%m-%d')
                        elif isinstance(tx_date, str) and tx_date != '-':
                            date_str = tx_date
                        transactions.append({
                            'date': date_str,
                            'name': tx_name,
                            'amount': float(tx_amount),
                            'category': tx_category
                        })
                monthly_data[month_key] = {
                    'categories': month_cats,
                    'transactions': transactions
                }
    return monthly_data

def load_data():
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}

def save_data(data):
    with open(DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

@app.route('/')
def index():
    return send_from_directory('static', 'index.html')

@app.route('/api/data')
def get_data():
    data = load_data()
    return jsonify(data)

@app.route('/api/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': '파일이 없습니다'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '파일명이 없습니다'}), 400
    if not file.filename.endswith('.xlsx'):
        return jsonify({'error': '.xlsx 파일만 업로드 가능합니다'}), 400
    filepath = os.path.join(UPLOAD_FOLDER, 'uploaded.xlsx')
    file.save(filepath)
    try:
        new_data = parse_excel(filepath)
        existing_data = load_data()
        # Merge: new data overrides existing for same months
        merged = {**existing_data, **new_data}
        save_data(merged)
        return jsonify({
            'success': True,
            'months_added': list(new_data.keys()),
            'total_months': len(merged)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/summary')
def get_summary():
    data = load_data()
    months = sorted(data.keys())
    summary = []
    for month in months:
        cats = data[month].get('categories', {})
        total = sum(cats.values())
        summary.append({
            'month': month,
            'total': total,
            'categories': cats
        })
    return jsonify(summary)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=False)
